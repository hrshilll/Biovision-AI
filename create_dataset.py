import cv2
import mediapipe as mp
import numpy as np
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR       = "/Users/harshilchauhan/Desktop/BioVision AI/Data"
OUTPUT_EXCEL   = "gym_dataset.xlsx"
OUTPUT_CSV     = "gym_dataset.csv"
MIN_VISIBILITY = 0.30
MOTION_THRESH  = 1.5  

EXCLUDED_DIRS  = {
    "__pycache__", ".git", ".venv", "venv", "env",
    ".DS_Store", "models", "reports",
    "debug_visibility.py", "create_dataset.py",
    "analyze_form.py"
}

# Supported exercises (used for validation messaging only)
SUPPORTED_EXERCISES = {"pushup", "plank", "bicep_curl", "squat", "deadlift"}

mp_pose = mp.solutions.pose

def calculate_angle(a, b, c) -> float:
    a, b, c = np.array(a), np.array(b), np.array(c)
    radians = np.arctan2(c[1]-b[1], c[0]-b[0]) - np.arctan2(a[1]-b[1], a[0]-b[0])
    angle = np.abs(np.degrees(radians))
    return round(360.0 - angle if angle > 180.0 else angle, 2)


def get_lm(landmarks, name, min_vis):
    lm = landmarks[mp_pose.PoseLandmark[name].value]
    if lm.visibility < min_vis:
        raise ValueError(f"Low visibility: {name} ({lm.visibility:.2f})")
    return [lm.x, lm.y]


def extract_angles(lms, side: str, min_vis: float) -> dict:
    s = side.upper()
    shoulder = get_lm(lms, f"{s}_SHOULDER", min_vis)
    elbow    = get_lm(lms, f"{s}_ELBOW",    min_vis)
    wrist    = get_lm(lms, f"{s}_WRIST",    min_vis)
    hip      = get_lm(lms, f"{s}_HIP",      min_vis)
    knee     = get_lm(lms, f"{s}_KNEE",     min_vis)
    ankle    = get_lm(lms, f"{s}_ANKLE",    min_vis)
    return {
        "Elbow_Angle"    : calculate_angle(shoulder, elbow, wrist),
        "Shoulder_Angle" : calculate_angle(hip, shoulder, elbow),
        "Hip_Angle"      : calculate_angle(shoulder, hip, knee),
        "Knee_Angle"     : calculate_angle(hip, knee, ankle),
    }


def is_static(prev, curr, threshold) -> bool:
    if prev is None:
        return False
    return all(abs(curr[k] - prev[k]) < threshold for k in curr)


HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
GOOD_FILL    = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL     = PatternFill("solid", fgColor="FFCCCC")
ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT    = Font(bold=True, size=10)
NORMAL_FONT  = Font(size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

COLUMNS = [
    ("Exercise",       18),
    ("Video_File",     38),
    ("Frame_Number",   14),
    ("Side_Used",      11),
    ("Elbow_Angle",    13),
    ("Shoulder_Angle", 15),
    ("Hip_Angle",      11),
    ("Knee_Angle",     11),
    ("Form_Label",     12),
    ("Form",           10),
]


def style_header(ws):
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def style_data_row(ws, row_idx, values, label):
    fill = GOOD_FILL if label == 1 else BAD_FILL
    alt  = ALT_FILL  if row_idx % 2 == 0 else None

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border    = THIN_BORDER
        cell.font      = NORMAL_FONT
       
        if col_idx in (9, 10):
            cell.fill      = fill
            cell.font      = Font(bold=True, size=10,
                                  color="375623" if label == 1 else "9C0006")
            cell.alignment = CENTER
        else:
            cell.fill      = alt if alt else PatternFill()
            cell.alignment = CENTER if col_idx in (1, 3, 4, 5, 6, 7, 8) else LEFT_ALIGN


def main():
    rows = []

    with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:

        exercise_dirs = sorted([
            d for d in os.listdir(BASE_DIR)
            if os.path.isdir(os.path.join(BASE_DIR, d))
            and not d.startswith(".")
            and d not in EXCLUDED_DIRS
        ])

        if not exercise_dirs:
            print("⚠️  No exercise folders found. Check BASE_DIR.")
            return

        print(f"📂  Found exercises: {exercise_dirs}\n")

        for exercise in exercise_dirs:
            exercise_path = os.path.join(BASE_DIR, exercise)

            for label_name, label_value in [("good", 1), ("bad", 0)]:
              
                matched_dir = next(
                    (d for d in os.listdir(exercise_path)
                     if d.strip().lower() == label_name
                     and os.path.isdir(os.path.join(exercise_path, d))),
                    None
                )
                if matched_dir is None:
                    print(f"   ⚠️  Missing: {exercise}/{label_name} — skipping.")
                    continue
                video_dir = os.path.join(exercise_path, matched_dir)

                video_files = sorted([
                    f for f in os.listdir(video_dir)
                    if f.lower().endswith((".mp4", ".mov", ".mov.mov"))
                    and not f.startswith(".")
                ])

                if not video_files:
                    print(f"   ⚠️  No videos in {video_dir} — skipping.")
                    continue

                for video_file in video_files:
                    video_path = os.path.join(video_dir, video_file)
                    print(f"🔄  {exercise:15s} | {label_name:4s} | {video_file}")

                    cap = cv2.VideoCapture(video_path)
                    if not cap.isOpened():
                        print(f"   ❌ Cannot open — skipping.")
                        continue

                    frame_idx      = 0
                    rows_written   = 0
                    skipped_static = 0
                    skipped_vis    = 0
                    prev_angles    = None

                    while True:
                        ret, frame = cap.read()
                        if not ret:
                            break

                        frame_idx += 1
                        image   = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        results = pose.process(image)

                        if not results.pose_landmarks:
                            continue

                        lms       = results.pose_landmarks.landmark
                        angles    = None
                        side_used = None

                        for side in ("LEFT", "RIGHT"):
                            try:
                                angles    = extract_angles(lms, side, MIN_VISIBILITY)
                                side_used = side
                                break
                            except ValueError:
                                continue

                        if angles is None:
                            skipped_vis += 1
                            continue

                        if is_static(prev_angles, angles, MOTION_THRESH):
                            skipped_static += 1
                            prev_angles = angles
                            continue

                        prev_angles = angles

                        rows.append({
                            "Exercise"      : exercise,
                            "Video_File"    : video_file,
                            "Frame_Number"  : frame_idx,
                            "Side_Used"     : side_used,
                            "Elbow_Angle"   : angles["Elbow_Angle"],
                            "Shoulder_Angle": angles["Shoulder_Angle"],
                            "Hip_Angle"     : angles["Hip_Angle"],
                            "Knee_Angle"    : angles["Knee_Angle"],
                            "Form_Label"    : label_value,
                            "Form"          : "Good" if label_value == 1 else "Bad",
                        })
                        rows_written += 1

                    cap.release()
                    print(
                        f"   ✅  {rows_written:4d} rows  |  "
                        f"{skipped_static:4d} static skipped  |  "
                        f"{skipped_vis:4d} low-vis skipped"
                    )

    if not rows:
        print("\n⚠️  No data collected. Check folder structure and video files.")
        return

    df = pd.DataFrame(rows)

    
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"\n💾  CSV saved  → {OUTPUT_CSV}  ({len(df)} rows)")

    
    wb = Workbook()

    
    ws_data = wb.active
    ws_data.title = "Full Dataset"
    style_header(ws_data)

    for i, row in enumerate(df.itertuples(index=False), start=2):
        values = [
            row.Exercise, row.Video_File, row.Frame_Number, row.Side_Used,
            row.Elbow_Angle, row.Shoulder_Angle, row.Hip_Angle, row.Knee_Angle,
            row.Form_Label, row.Form,
        ]
        style_data_row(ws_data, i, values, row.Form_Label)

   
    ws_sum = wb.create_sheet("Summary")
    sum_header_fill = PatternFill("solid", fgColor="1F4E79")

    summary_cols = [
        ("Exercise", 18), ("Form", 10), ("Total Frames", 14),
        ("Avg Elbow°", 12), ("Avg Shoulder°", 14),
        ("Avg Hip°", 10), ("Avg Knee°", 10),
    ]
    for col_idx, (name, width) in enumerate(summary_cols, start=1):
        cell = ws_sum.cell(row=1, column=col_idx, value=name)
        cell.fill      = sum_header_fill
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = width
    ws_sum.row_dimensions[1].height = 22
    ws_sum.freeze_panes = "A2"

    summary = (
        df.groupby(["Exercise", "Form"])
        .agg(
            Total_Frames   = ("Frame_Number", "count"),
            Avg_Elbow      = ("Elbow_Angle",    "mean"),
            Avg_Shoulder   = ("Shoulder_Angle", "mean"),
            Avg_Hip        = ("Hip_Angle",       "mean"),
            Avg_Knee       = ("Knee_Angle",      "mean"),
        )
        .round(1)
        .reset_index()
    )

    for i, row in enumerate(summary.itertuples(index=False), start=2):
        label = 1 if row.Form == "Good" else 0
        fill  = GOOD_FILL if label == 1 else BAD_FILL
        vals  = [row.Exercise, row.Form, row.Total_Frames,
                 row.Avg_Elbow, row.Avg_Shoulder, row.Avg_Hip, row.Avg_Knee]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws_sum.cell(row=i, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = Font(bold=True, size=10,
                                  color="375623" if label == 1 else "9C0006")
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

    # ── Sheet 3: Per-Exercise sheets ──────────
    for exercise in df["Exercise"].unique():
        ws_ex = wb.create_sheet(exercise[:31])   # sheet name max 31 chars
        style_header(ws_ex)
        ex_df = df[df["Exercise"] == exercise]
        for i, row in enumerate(ex_df.itertuples(index=False), start=2):
            values = [
                row.Exercise, row.Video_File, row.Frame_Number, row.Side_Used,
                row.Elbow_Angle, row.Shoulder_Angle, row.Hip_Angle, row.Knee_Angle,
                row.Form_Label, row.Form,
            ]
            style_data_row(ws_ex, i, values, row.Form_Label)

    wb.save(OUTPUT_EXCEL)
    print(f"📊  Excel saved → {OUTPUT_EXCEL}")

    print("\n" + "="*60)
    print("  DATASET SUMMARY")
    print("="*60)
    print(df.groupby(["Exercise", "Form"])
            .size()
            .rename("Frames")
            .to_string())
    print(f"\n  Total rows : {len(df)}")
    print("="*60)


if __name__ == "__main__":
    main()
