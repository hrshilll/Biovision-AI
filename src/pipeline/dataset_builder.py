"""Video-to-dataset extraction pipeline."""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, List, Optional

import cv2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.config import load_settings, resolve_data_dir
from src.utils.logging_config import setup_logging
from src.utils.pose_extractor import PoseProcessor

logger = setup_logging(__name__)

EXCLUDED_DIRS = {
    "__pycache__", ".git", ".venv", "venv", "env",
    ".DS_Store", "models", "reports", "session_results", "tests",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFCCCC")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    ("Exercise", 18),
    ("Video_File", 38),
    ("Frame_Number", 14),
    ("Side_Used", 11),
    ("Elbow_Angle", 13),
    ("Shoulder_Angle", 15),
    ("Hip_Angle", 11),
    ("Knee_Angle", 11),
    ("Elbow_Flexion_3D", 15),
    ("Knee_Flexion_3D", 15),
    ("Hip_Flexion_3D", 14),
    ("Shoulder_Abduction_3D", 18),
    ("Shoulder_Flexion_3D", 17),
    ("Trunk_Inclination_3D", 18),
    ("Spine_Alignment_3D", 18),
    ("Pelvic_Tilt_3D", 14),
    ("Form_Label", 12),
    ("Form", 10),
]


def is_static(prev: Optional[Dict[str, float]], curr: Dict[str, float], threshold: float) -> bool:
    if prev is None:
        return False
    keys = ["Elbow_Angle", "Shoulder_Angle", "Hip_Angle", "Knee_Angle"]
    return all(abs(curr.get(k, 0) - prev.get(k, 0)) < threshold for k in keys)


def style_header(ws) -> None:
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def style_data_row(ws, row_idx: int, values: List[Any], label: int) -> None:
    fill = GOOD_FILL if label == 1 else BAD_FILL
    alt = ALT_FILL if row_idx % 2 == 0 else None
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.font = NORMAL_FONT
        if col_idx >= len(COLUMNS) - 1:
            cell.fill = fill
            cell.font = Font(
                bold=True, size=10, color="375623" if label == 1 else "9C0006"
            )
            cell.alignment = CENTER
        else:
            cell.fill = alt if alt else PatternFill()
            cell.alignment = CENTER if col_idx not in (1, 2) else LEFT_ALIGN


def build_dataset(
    base_dir: Optional[Path] = None,
    output_csv: str = "gym_dataset.csv",
    output_excel: str = "gym_dataset.xlsx",
) -> pd.DataFrame:
    settings = load_settings()
    pose_cfg = settings.get("pose", {})
    base = base_dir or resolve_data_dir(settings)
    min_vis = float(pose_cfg.get("min_visibility", 0.30))
    motion_thresh = float(pose_cfg.get("motion_threshold", 1.5))

    rows: List[Dict[str, Any]] = []
    if not base.exists():
        logger.error("Data directory not found: %s", base)
        return pd.DataFrame()

    with PoseProcessor(
        min_detection_confidence=float(pose_cfg.get("min_detection_confidence", 0.5)),
        min_tracking_confidence=float(pose_cfg.get("min_tracking_confidence", 0.5)),
        model_complexity=int(pose_cfg.get("model_complexity", 1)),
        min_visibility=min_vis,
        use_3d=True,
        include_posture=True,
    ) as pose:
        exercise_dirs = sorted(
            d for d in os.listdir(base)
            if (base / d).is_dir() and not d.startswith(".") and d not in EXCLUDED_DIRS
        )
        if not exercise_dirs:
            logger.warning("No exercise folders found in %s", base)
            return pd.DataFrame()

        logger.info("Found exercises: %s", exercise_dirs)
        for exercise in exercise_dirs:
            exercise_path = base / exercise
            for label_name, label_value in [("good", 1), ("bad", 0)]:
                matched = next(
                    (
                        d for d in os.listdir(exercise_path)
                        if d.strip().lower() == label_name and (exercise_path / d).is_dir()
                    ),
                    None,
                )
                if matched is None:
                    logger.warning("Missing folder: %s/%s", exercise, label_name)
                    continue
                video_dir = exercise_path / matched
                video_files = sorted(
                    f for f in os.listdir(video_dir)
                    if f.lower().endswith((".mp4", ".mov", ".mov.mov")) and not f.startswith(".")
                )
                for video_file in video_files:
                    video_path = video_dir / video_file
                    logger.info("Processing %s | %s | %s", exercise, label_name, video_file)
                    cap = cv2.VideoCapture(str(video_path))
                    if not cap.isOpened():
                        logger.error("Cannot open video: %s", video_path)
                        continue
                    frame_idx = 0
                    prev_angles = None
                    while True:
                        ret, frame = cap.read()
                        if not ret:
                            break
                        frame_idx += 1
                        result = pose.process_frame(frame)
                        if result is None:
                            continue
                        angles = result.angles
                        if is_static(prev_angles, angles, motion_thresh):
                            prev_angles = angles
                            continue
                        prev_angles = angles
                        rows.append(
                            {
                                "Exercise": exercise,
                                "Video_File": video_file,
                                "Frame_Number": frame_idx,
                                "Side_Used": result.side_used,
                                "Elbow_Angle": angles.get("Elbow_Angle"),
                                "Shoulder_Angle": angles.get("Shoulder_Angle"),
                                "Hip_Angle": angles.get("Hip_Angle"),
                                "Knee_Angle": angles.get("Knee_Angle"),
                                "Elbow_Flexion_3D": angles.get("Elbow_Flexion_3D"),
                                "Knee_Flexion_3D": angles.get("Knee_Flexion_3D"),
                                "Hip_Flexion_3D": angles.get("Hip_Flexion_3D"),
                                "Shoulder_Abduction_3D": angles.get("Shoulder_Abduction_3D"),
                                "Shoulder_Flexion_3D": angles.get("Shoulder_Flexion_3D"),
                                "Trunk_Inclination_3D": angles.get("Trunk_Inclination_3D"),
                                "Spine_Alignment_3D": angles.get("Spine_Alignment_3D"),
                                "Pelvic_Tilt_3D": angles.get("Pelvic_Tilt_3D"),
                                "Form_Label": label_value,
                                "Form": "Good" if label_value == 1 else "Bad",
                            }
                        )
                    cap.release()

    if not rows:
        logger.warning("No data collected.")
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df.to_csv(output_csv, index=False)
    logger.info("CSV saved → %s (%d rows)", output_csv, len(df))

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Full Dataset"
    style_header(ws_data)
    col_names = [c[0] for c in COLUMNS]
    for i, row in enumerate(df.itertuples(index=False), start=2):
        values = [getattr(row, name) for name in col_names]
        style_data_row(ws_data, i, values, row.Form_Label)
    wb.save(output_excel)
    logger.info("Excel saved → %s", output_excel)
    return df
