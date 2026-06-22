"""Session analytics and export utilities."""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Mapping, Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.biomechanics.kinematics import SessionKinematics
from src.exercises.form_scorer import FormScore

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class SessionAnalytics:
    """Aggregated workout session analytics."""

    exercise: str
    total_frames: int
    reps: int
    form_score: FormScore
    kinematics: SessionKinematics
    error_counts: Dict[str, int] = field(default_factory=dict)
    error_messages: List[str] = field(default_factory=list)

    def to_summary_dict(self) -> Dict[str, float]:
        summary = {
            "Exercise": self.exercise,
            "Total_Frames": self.total_frames,
            "Reps": self.reps,
            **self.form_score.to_dict(),
            "Stability": self.kinematics.stability_score,
            "Symmetry": self.kinematics.symmetry_score,
            "Smoothness": self.kinematics.smoothness_score,
            "Trajectory_Consistency": self.kinematics.trajectory_consistency,
        }
        for key, value in self.kinematics.rom.items():
            summary[f"ROM_{key}"] = value
        for code, count in self.error_counts.items():
            summary[f"Error_{code}"] = count
        return summary


class SessionAnalyzer:
    """Build analytics from a session log."""

    def analyze(
        self,
        session_log: List[Mapping],
        exercise: str,
        form_score: FormScore,
        kinematics: SessionKinematics,
    ) -> SessionAnalytics:
        error_counts: Dict[str, int] = {}
        error_messages: List[str] = []
        for row in session_log:
            issues_text = str(row.get("Issues", "") or row.get("Error_Codes", ""))
            if not issues_text:
                continue
            for part in issues_text.split(";"):
                msg = part.strip()
                if not msg:
                    continue
                error_messages.append(msg)
                code = msg.split("—")[0].strip().lower().replace(" ", "_")[:40]
                error_counts[code] = error_counts.get(code, 0) + 1

        reps = 0
        if session_log:
            reps = int(session_log[-1].get("Reps", 0))

        return SessionAnalytics(
            exercise=exercise,
            total_frames=len(session_log),
            reps=reps,
            form_score=form_score,
            kinematics=kinematics,
            error_counts=error_counts,
            error_messages=sorted(set(error_messages)),
        )


def _style_sheet(ws, freeze: str = "A2") -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        header = str(cell.value or "")
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 12), 28)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = freeze

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN if col_idx > 2 else CENTER
            cell.border = THIN_BORDER


def export_session_csv(df: pd.DataFrame, path: str) -> None:
    df.to_csv(path, index=False)


def export_session_excel(
    df: pd.DataFrame,
    path: str,
    summary: Optional[Mapping] = None,
    landmarks_df: Optional[pd.DataFrame] = None,
) -> None:
    """Export session data to a formatted Excel workbook."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Session", index=False)
        if summary:
            pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)

        if landmarks_df is not None and not landmarks_df.empty:
            landmarks_df.to_excel(writer, sheet_name="Landmarks", index=False)

        if "Reps" in df.columns:
            rep_changes = df[df["Reps"].diff().fillna(0) > 0][["Frame", "Reps", "Exercise"]].copy()
            if not rep_changes.empty:
                rep_changes.to_excel(writer, sheet_name="Rep_Events", index=False)

        for sheet_name in writer.sheets:
            _style_sheet(writer.sheets[sheet_name])


def export_analytics_report(analytics: SessionAnalytics, txt_path: str) -> None:
    lines = [
        "=" * 72,
        "  BIOVISION AI — SESSION ANALYTICS REPORT",
        "=" * 72,
        f"Exercise     : {analytics.exercise}",
        f"Frames       : {analytics.total_frames}",
        f"Reps         : {analytics.reps}",
        "",
        analytics.form_score.formatted(),
        "",
        f"Stability    : {analytics.kinematics.stability_score}",
        f"Symmetry     : {analytics.kinematics.symmetry_score}",
        f"Smoothness   : {analytics.kinematics.smoothness_score}",
        f"Tempo        : {analytics.kinematics.tempo_score}",
        "",
        "Range of Motion:",
    ]
    for key, val in analytics.kinematics.rom.items():
        lines.append(f"  {key}: {val}°")
    if analytics.error_counts:
        lines.append("\nError Frequency:")
        for code, count in sorted(analytics.error_counts.items(), key=lambda x: -x[1]):
            lines.append(f"  {code}: {count}")
    lines.append("=" * 72)
    with open(txt_path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
