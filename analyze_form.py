#!/usr/bin/env python3
"""Offline form analysis report from gym_dataset.csv."""

from __future__ import annotations

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.exercises.config_loader import ExerciseRegistry
from src.exercises.form_checker import check_form
from src.utils.config import get_path
from src.utils.logging_config import setup_logging

OUTPUT_EXCEL = "form_analysis_report.xlsx"
OUTPUT_TXT = "form_analysis_report.txt"

H_FILL = PatternFill("solid", fgColor="1F4E79")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFCCCC")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def hcell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = H_FILL
    c.font = H_FONT
    c.alignment = CENTER
    c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def dcell(ws, row, col, value, fill=None, bold=False, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.font = Font(bold=bold, size=10)
    c.alignment = align
    c.border = THIN
    return c


def analyse_group(df_video: pd.DataFrame, exercise: str, registry: ExerciseRegistry) -> dict:
    issues = {}
    cfg = registry.get(exercise)
    if not cfg:
        return issues
    for _, row in df_video.iterrows():
        angles = row.to_dict()
        _, form_issues = check_form(angles, exercise, registry=registry)
        for issue in form_issues:
            key = issue.metric or issue.code
            if key not in issues:
                issues[key] = {
                    "reason": issue.message,
                    "range": (issue.lo, issue.hi),
                    "bad_count": 0,
                    "total": len(df_video),
                    "values": [],
                }
            issues[key]["bad_count"] += 1
            issues[key]["values"].append(float(issue.value or 0))
    for key, info in issues.items():
        info["pct"] = round(100 * info["bad_count"] / max(info["total"], 1), 1)
        info["mean"] = round(float(sum(info["values"]) / max(len(info["values"]), 1)), 1)
    return issues


def build_excel(df: pd.DataFrame, registry: ExerciseRegistry) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.freeze_panes = "A2"
    ov_cols = [
        ("Exercise", 18), ("Video File", 38), ("Form", 8),
        ("Frames Analysed", 16), ("Issues Found", 14), ("Verdict", 40),
    ]
    for i, (name, w) in enumerate(ov_cols, 1):
        hcell(ws, 1, i, name, w)

    row_idx = 2
    for (exercise, video_file, form_label), grp in df.groupby(["Exercise", "Video_File", "Form_Label"]):
        issues = analyse_group(grp, exercise, registry)
        verdict = "Good Form" if form_label == 1 else (
            "Bad Form — " + "; ".join(i["reason"] for i in issues.values()) if issues else "Bad Form"
        )
        fill = GOOD_FILL if form_label == 1 else BAD_FILL
        vals = [exercise, video_file, "Good" if form_label == 1 else "Bad", len(grp), len(issues), verdict]
        for col, val in enumerate(vals, 1):
            dcell(ws, row_idx, col, val, fill=fill, align=LEFT if col == 6 else CENTER)
        row_idx += 1
    wb.save(OUTPUT_EXCEL)


def build_text(df: pd.DataFrame, registry: ExerciseRegistry) -> str:
    lines = ["=" * 72, "  BIOVISION AI — FORM ANALYSIS REPORT", "=" * 72]
    for (exercise, video_file, form_label), grp in df.groupby(["Exercise", "Video_File", "Form_Label"]):
        lines.extend(
            [
                f"\nExercise  : {exercise}",
                f"Video     : {video_file}",
                f"Verdict   : {'GOOD FORM' if form_label == 1 else 'BAD FORM'}",
                f"Frames    : {len(grp)}",
            ]
        )
        issues = analyse_group(grp, exercise, registry)
        if not issues:
            lines.append("  → All joint angles within acceptable range.")
        else:
            lines.append("\n  Issues:")
            for angle, info in issues.items():
                lines.append(f"  [{angle}] {info['reason']} ({info['pct']}% frames)")
        lines.append("-" * 72)
    return "\n".join(lines)


def main() -> None:
    logger = setup_logging(__name__)
    input_csv = str(get_path("dataset_csv", "gym_dataset.csv"))
    registry = ExerciseRegistry()
    try:
        df = pd.read_csv(input_csv)
    except FileNotFoundError:
        logger.error("%s not found. Run create_dataset.py first.", input_csv)
        return
    if df.empty:
        logger.error("Dataset is empty.")
        return
    build_excel(df, registry)
    report = build_text(df, registry)
    print(report)
    with open(OUTPUT_TXT, "w", encoding="utf-8") as handle:
        handle.write(report)
    logger.info("Reports saved → %s, %s", OUTPUT_EXCEL, OUTPUT_TXT)


if __name__ == "__main__":
    main()
