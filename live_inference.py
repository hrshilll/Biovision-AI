"""
live_inference.py  —  BioVision AI
────────────────────────────────────
Flow:
  1. Terminal asks which exercise to perform
  2. Webcam opens with live skeleton + angle overlay
  3. Reps counted accurately per exercise using angle state machine
  4. Form verdict (Good/Bad) shown per frame with specific issue messages
  5. Press SPACE or Q to stop — session saved to session_results/
"""

import cv2
import mediapipe as mp
import numpy as np
import pickle
import os
import time
import datetime
import pandas as pd
from collections import deque

# ── Paths ──────────────────────────────────────────────────────────
MODEL_DIR    = "models"
SESSION_DIR  = "session_results"
os.makedirs(SESSION_DIR, exist_ok=True)

# ── Tuning ─────────────────────────────────────────────────────────
MIN_VIS          = 0.30
SMOOTH_N         = 8      # prediction smoothing window

# ── Exercise definitions ────────────────────────────────────────────
EXERCISES = {
    "1": "Bicep_curls",
    "2": "Deadlifts",
    "3": "Planks",
    "4": "Pushups",
    "5": "Squats",
}

# ── Biomechanical rules (lo, hi, short_reason) ─────────────────────
RULES = {
    "Bicep_curls": {
        "Elbow_Angle"    : (20,  160, "Incomplete elbow range of motion"),
        "Shoulder_Angle" : (0,   40,  "Upper arm swinging — using momentum"),
        "Hip_Angle"      : (155, 180, "Leaning back — using body swing"),
        "Knee_Angle"     : (155, 180, "Knees bending — unstable base"),
    },
    "Pushups": {
        "Elbow_Angle"    : (55,  175, "Elbow not reaching 90 or not extending"),
        "Shoulder_Angle" : (25,  95,  "Shoulders flaring or collapsing"),
        "Hip_Angle"      : (155, 180, "Hips sagging or piking"),
        "Knee_Angle"     : (155, 180, "Knees bent — not in full plank"),
    },
    "Planks": {
        "Elbow_Angle"    : (75,  105, "Elbows not at 90 forearm position"),
        "Shoulder_Angle" : (75,  105, "Shoulders not over elbows"),
        "Hip_Angle"      : (155, 180, "Hips sagging or piking"),
        "Knee_Angle"     : (155, 180, "Knees bent — legs must be straight"),
    },
    "Squats": {
        "Elbow_Angle"    : (55,  180, "Arms not in stable position"),
        "Shoulder_Angle" : (55,  180, "Excessive forward lean"),
        "Hip_Angle"      : (45,  125, "Not reaching depth or spine collapse"),
        "Knee_Angle"     : (55,  125, "Knees not bending or caving in"),
    },
    "Deadlifts": {
        "Elbow_Angle"    : (155, 180, "Arms bent — should be straight"),
        "Shoulder_Angle" : (55,  180, "Shoulders rounding — injury risk"),
        "Hip_Angle"      : (55,  165, "Hips not hinging properly"),
        "Knee_Angle"     : (95,  175, "Knee angle incorrect at start/finish"),
    },
}

# ── Rep counting config per exercise ───────────────────────────────
# (angle_key, down_threshold, up_threshold, count_on)
# count_on: "up" = count when going from down→up, "down" = down→up→down
REP_CONFIG = {
    "Bicep_curls": ("Elbow_Angle",    155, 50,  "up"),
    "Pushups"    : ("Elbow_Angle",    160, 85,  "down"),
    "Squats"     : ("Knee_Angle",     165, 85,  "down"),
    "Deadlifts"  : ("Hip_Angle",      160, 80,  "up"),
    "Planks"     : None,   # isometric — no reps
}

# ── Colours (BGR) ──────────────────────────────────────────────────
GREEN  = (50,  205, 50)
RED    = (30,  30,  220)
ORANGE = (0,   140, 255)
WHITE  = (255, 255, 255)
BLACK  = (0,   0,   0)
DGREY  = (35,  35,  35)
LGREY  = (170, 170, 170)
YELLOW = (0,   215, 255)
CYAN   = (255, 200, 0)
TEAL   = (180, 180, 0)

mp_pose    = mp.solutions.pose
mp_drawing = mp.solutions.drawing_utils


# ══════════════════════════════════════════════
# GEOMETRY
# ══════════════════════════════════════════════
def calc_angle(a, b, c):
    a, b, c = np.array(a), np.array(b), np.array(c)
    r = np.arctan2(c[1]-b[1], c[0]-b[0]) - np.arctan2(a[1]-b[1], a[0]-b[0])
    ang = np.abs(np.degrees(r))
    return round(360 - ang if ang > 180 else ang, 1)


def get_lm(lms, name):
    lm = lms[mp_pose.PoseLandmark[name].value]
    if lm.visibility < MIN_VIS:
        raise ValueError(name)
    return [lm.x, lm.y]


def get_lm_px(lms, name, w, h):
    lm = lms[mp_pose.PoseLandmark[name].value]
    return int(lm.x * w), int(lm.y * h)


def extract_angles(lms, side):
    s = side.upper()
    sh = get_lm(lms, f"{s}_SHOULDER")
    el = get_lm(lms, f"{s}_ELBOW")
    wr = get_lm(lms, f"{s}_WRIST")
    hi = get_lm(lms, f"{s}_HIP")
    kn = get_lm(lms, f"{s}_KNEE")
    an = get_lm(lms, f"{s}_ANKLE")
    return {
        "Elbow_Angle"    : calc_angle(sh, el, wr),
        "Shoulder_Angle" : calc_angle(hi, sh, el),
        "Hip_Angle"      : calc_angle(sh, hi, kn),
        "Knee_Angle"     : calc_angle(hi, kn, an),
    }, side


def best_angles(lms):
    for side in ("LEFT", "RIGHT"):
        try:
            return extract_angles(lms, side)
        except ValueError:
            continue
    return None, None


# ══════════════════════════════════════════════
# REP COUNTER  — proper state machine
# ══════════════════════════════════════════════
class RepCounter:
    def __init__(self, exercise):
        self.exercise = exercise
        self.count    = 0
        self.stage    = "up"   # "up" or "down"
        self.cfg      = REP_CONFIG.get(exercise)

    def update(self, angles):
        if self.cfg is None:
            return   # isometric exercise
        angle_key, down_thr, up_thr, count_on = self.cfg
        val = angles.get(angle_key, 0)

        if val > down_thr:
            new_stage = "up"
        elif val < up_thr:
            new_stage = "down"
        else:
            return   # in transition zone — don't change stage

        if new_stage != self.stage:
            # Stage just changed
            if count_on == "up"   and new_stage == "up":
                self.count += 1
            elif count_on == "down" and new_stage == "down":
                self.count += 1
            self.stage = new_stage

    def reset(self):
        self.count = 0
        self.stage = "up"


# ══════════════════════════════════════════════
# FORM CHECKER
# ══════════════════════════════════════════════
def check_form(angles, exercise):
    """Returns (is_good, issues_list).
    issues_list: [(angle_name, reason, val, lo, hi), ...]
    """
    rules  = RULES.get(exercise, {})
    issues = []
    for angle, (lo, hi, reason) in rules.items():
        val = angles.get(angle, 0)
        if not (lo <= val <= hi):
            issues.append((angle, reason, val, lo, hi))
    return len(issues) == 0, issues


# ══════════════════════════════════════════════
# DRAWING HELPERS
# ══════════════════════════════════════════════
def txt(img, text, pos, scale=0.55, color=WHITE, bold=False):
    t = 2 if bold else 1
    cv2.putText(img, text, pos, cv2.FONT_HERSHEY_SIMPLEX, scale, BLACK, t+2, cv2.LINE_AA)
    cv2.putText(img, text, pos, cv2.FONT_HERSHEY_SIMPLEX, scale, color, t,   cv2.LINE_AA)


def panel(img, x1, y1, x2, y2, color=DGREY, alpha=0.82):
    sub = img[y1:y2, x1:x2]
    rect = np.full_like(sub, color)
    cv2.addWeighted(rect, alpha, sub, 1-alpha, 0, sub)
    img[y1:y2, x1:x2] = sub


def bar(img, x, y, w, h, pct, fg, bg=DGREY):
    cv2.rectangle(img, (x, y), (x+w, y+h), bg, -1)
    cv2.rectangle(img, (x, y), (x+int(w*pct), y+h), fg, -1)


def draw_angle_on_joint(img, lms, side, angle_name, val, w, h, color):
    joint_map = {
        "Elbow_Angle"    : f"{side.upper()}_ELBOW",
        "Shoulder_Angle" : f"{side.upper()}_SHOULDER",
        "Hip_Angle"      : f"{side.upper()}_HIP",
        "Knee_Angle"     : f"{side.upper()}_KNEE",
    }
    jname = joint_map.get(angle_name)
    if not jname:
        return
    lm = lms[mp_pose.PoseLandmark[jname].value]
    if lm.visibility < MIN_VIS:
        return
    cx, cy = int(lm.x * w), int(lm.y * h)
    cv2.circle(img, (cx, cy), 12, color, -1)
    cv2.circle(img, (cx, cy), 12, WHITE,  1)
    txt(img, f"{val:.0f}", (cx-14, cy+5), scale=0.42, color=WHITE, bold=True)


def draw_skeleton_colored(img, lms, angles, exercise, w, h):
    """Draw skeleton with joints coloured green/red based on form."""
    rules = RULES.get(exercise, {})
    joint_angle_map = {
        "LEFT_ELBOW"    : "Elbow_Angle",
        "RIGHT_ELBOW"   : "Elbow_Angle",
        "LEFT_SHOULDER" : "Shoulder_Angle",
        "RIGHT_SHOULDER": "Shoulder_Angle",
        "LEFT_HIP"      : "Hip_Angle",
        "RIGHT_HIP"     : "Hip_Angle",
        "LEFT_KNEE"     : "Knee_Angle",
        "RIGHT_KNEE"    : "Knee_Angle",
    }

    # Draw connections
    for conn in mp_pose.POSE_CONNECTIONS:
        a_idx, b_idx = conn
        a_lm = lms[a_idx]
        b_lm = lms[b_idx]
        if a_lm.visibility < MIN_VIS or b_lm.visibility < MIN_VIS:
            continue
        ax, ay = int(a_lm.x * w), int(a_lm.y * h)
        bx, by = int(b_lm.x * w), int(b_lm.y * h)
        cv2.line(img, (ax, ay), (bx, by), (120, 120, 120), 2, cv2.LINE_AA)

    # Draw joints
    for idx, lm in enumerate(lms):
        if lm.visibility < MIN_VIS:
            continue
        name = mp_pose.PoseLandmark(idx).name
        angle_key = joint_angle_map.get(name)
        color = LGREY
        if angle_key and angle_key in rules and angle_key in angles:
            lo, hi, _ = rules[angle_key]
            color = GREEN if lo <= angles[angle_key] <= hi else RED
        cx, cy = int(lm.x * w), int(lm.y * h)
        cv2.circle(img, (cx, cy), 5, color, -1)
        cv2.circle(img, (cx, cy), 5, WHITE, 1)


def draw_left_panel(img, exercise, is_good, confidence, issues,
                    angles, rep_counter, fps, h, side_used, lms, w):
    PW = 310
    panel(img, 0, 0, PW, h, DGREY, 0.88)
    cv2.line(img, (PW, 0), (PW, h), (80, 80, 80), 1)

    # ── Verdict header ──────────────────────
    hcol = GREEN if is_good else RED
    cv2.rectangle(img, (0, 0), (PW, 58), hcol, -1)
    verdict = "GOOD FORM" if is_good else "BAD FORM"
    txt(img, verdict, (12, 40), scale=1.0, color=WHITE, bold=True)

    # ── Exercise ────────────────────────────
    txt(img, exercise.replace("_", " ").upper(),
        (12, 80), scale=0.55, color=YELLOW, bold=True)

    # ── Confidence ──────────────────────────
    txt(img, f"Confidence: {confidence*100:.0f}%", (12, 106), scale=0.46, color=LGREY)
    bar(img, 12, 112, PW-24, 7, confidence, hcol)

    # ── Angles ──────────────────────────────
    y = 138
    txt(img, "JOINT ANGLES", (12, y), scale=0.44, color=CYAN)
    y += 4
    cv2.line(img, (12, y), (PW-12, y), (70, 70, 70), 1)
    y += 14
    rules = RULES.get(exercise, {})
    for aname, val in (angles or {}).items():
        lo, hi, _ = rules.get(aname, (0, 360, ""))
        ok    = lo <= val <= hi
        acol  = GREEN if ok else RED
        label = aname.replace("_Angle", "").replace("_", " ")
        # angle bar
        norm  = min(max((val - lo) / max(hi - lo, 1), 0), 1) if ok else 0
        bar(img, 12, y, PW-24, 5, norm if ok else 1.0,
            GREEN if ok else RED, (60, 60, 60))
        txt(img, f"{label:<10} {val:>6.1f}", (12, y+18), scale=0.46, color=acol)
        y += 28

    # ── Rep counter ─────────────────────────
    y += 6
    cv2.rectangle(img, (10, y), (PW-10, y+62), (50, 50, 50), -1)
    cv2.rectangle(img, (10, y), (PW-10, y+62), (90, 90, 90), 1)
    cfg = REP_CONFIG.get(exercise)
    if cfg is None:
        txt(img, "ISOMETRIC HOLD", (20, y+22), scale=0.5, color=LGREY)
        txt(img, "No rep counting", (20, y+44), scale=0.42, color=LGREY)
    else:
        txt(img, "REPS", (20, y+20), scale=0.5, color=LGREY)
        txt(img, str(rep_counter.count), (20, y+52),
            scale=1.5, color=WHITE, bold=True)
        stage_col = CYAN if rep_counter.stage == "down" else YELLOW
        txt(img, rep_counter.stage.upper(), (110, y+52),
            scale=0.75, color=stage_col, bold=True)
    y += 72

    # ── Issues ──────────────────────────────
    if issues:
        txt(img, "FORM ISSUES", (12, y+16), scale=0.46, color=ORANGE, bold=True)
        y += 22
        cv2.line(img, (12, y), (PW-12, y), (70, 70, 70), 1)
        y += 10
        for _, reason, val, lo, hi in issues[:4]:
            dev = min(abs(val-lo), abs(val-hi))
            short = reason[:34] + ".." if len(reason) > 36 else reason
            txt(img, f"• {short}", (12, y), scale=0.40, color=ORANGE)
            txt(img, f"  ({val:.0f} vs {lo}-{hi}, off {dev:.0f})",
                (12, y+14), scale=0.36, color=(120, 120, 180))
            y += 32
    else:
        txt(img, "All angles in range", (12, y+16), scale=0.46, color=GREEN)

    # ── FPS + controls ──────────────────────
    txt(img, f"FPS {fps:.0f}", (12, h-30), scale=0.42, color=LGREY)
    txt(img, "SPACE/Q = Stop & Save", (12, h-12), scale=0.40, color=LGREY)

    # ── Angle labels on skeleton ─────────────
    if angles and side_used and lms:
        for aname, val in angles.items():
            lo, hi, _ = rules.get(aname, (0, 360, ""))
            acol = GREEN if lo <= val <= hi else RED
            draw_angle_on_joint(img, lms, side_used, aname, val, w, h, acol)


def draw_exercise_menu(frame, h, w):
    """Full-screen exercise selection overlay."""
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, 0), (w, h), (15, 15, 15), -1)
    cv2.addWeighted(overlay, 0.92, frame, 0.08, 0, frame)

    cx = w // 2
    txt(frame, "BioVision AI", (cx-160, 80), scale=1.2, color=CYAN, bold=True)
    txt(frame, "Select Exercise to Begin", (cx-190, 120), scale=0.7, color=LGREY)
    cv2.line(frame, (cx-200, 135), (cx+200, 135), (80, 80, 80), 1)

    for key, name in EXERCISES.items():
        y = 160 + int(key) * 65
        bx1, by1 = cx-200, y-30
        bx2, by2 = cx+200, y+20
        cv2.rectangle(frame, (bx1, by1), (bx2, by2), (40, 40, 40), -1)
        cv2.rectangle(frame, (bx1, by1), (bx2, by2), (90, 90, 90), 1)
        txt(frame, f"[{key}]  {name.replace('_', ' ')}",
            (bx1+20, y+6), scale=0.65, color=WHITE, bold=True)

    txt(frame, "Press Q to quit", (cx-90, h-30), scale=0.45, color=LGREY)


# ══════════════════════════════════════════════
# LOAD MODEL
# ══════════════════════════════════════════════
def load_models():
    paths = {
        "clf"    : os.path.join(MODEL_DIR, "form_classifier.pkl"),
        "encoder": os.path.join(MODEL_DIR, "label_encoder.pkl"),
        "scaler" : os.path.join(MODEL_DIR, "scaler.pkl"),
    }
    missing = [k for k, p in paths.items() if not os.path.exists(p)]
    if missing:
        print(f"Missing model files: {missing}. Run train_model.py first.")
        return None, None, None
    with open(paths["clf"],     "rb") as f: clf     = __import__("pickle").load(f)
    with open(paths["encoder"], "rb") as f: encoder = __import__("pickle").load(f)
    with open(paths["scaler"],  "rb") as f: scaler  = __import__("pickle").load(f)
    return clf, encoder, scaler


# ══════════════════════════════════════════════
# SAVE SESSION
# ══════════════════════════════════════════════
def save_session(session_log, exercise):
    if not session_log:
        print("No data to save.")
        return
    ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    df  = pd.DataFrame(session_log)
    base = os.path.join(SESSION_DIR, f"{exercise}_{ts}")
    df.to_csv(base + ".csv", index=False)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Session"
        headers = list(df.columns)
        hfill = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[__import__("openpyxl.utils", fromlist=["get_column_letter"]).get_column_letter(ci)].width = 16
        gfill = PatternFill("solid", fgColor="C6EFCE")
        bfill = PatternFill("solid", fgColor="FFCCCC")
        for ri, row in enumerate(df.itertuples(index=False), start=2):
            is_good = row.Form == "Good"
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                if ci == headers.index("Form") + 1:
                    c.fill = gfill if is_good else bfill
                    c.font = Font(bold=True,
                                  color="375623" if is_good else "9C0006")
                c.alignment = Alignment(horizontal="center")
        wb.save(base + ".xlsx")
        print(f"Session saved: {base}.csv + .xlsx  ({len(df)} frames)")
    except Exception as e:
        print(f"Session saved: {base}.csv  ({len(df)} frames)  [xlsx skipped: {e}]")


# ══════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════
def main():
    clf, encoder, scaler = load_models()
    if clf is None:
        return

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Cannot open webcam. Grant camera permission and retry.")
        return
    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

    current_exercise = None
    rep_counter      = None
    pred_buf         = deque(maxlen=SMOOTH_N)
    conf_buf         = deque(maxlen=SMOOTH_N)
    session_log      = []
    prev_time        = time.time()
    frame_n          = 0
    screenshot_n     = 0
    os.makedirs("screenshots", exist_ok=True)

    print("\nBioVision AI started. Select exercise in the webcam window.")

    with mp_pose.Pose(min_detection_confidence=0.5,
                      min_tracking_confidence=0.5,
                      model_complexity=1) as pose:
        while True:
            ret, frame = cap.read()
            if not ret:
                break

            frame = cv2.flip(frame, 1)
            h, w  = frame.shape[:2]

            # ── Exercise selection screen ──────
            if current_exercise is None:
                draw_exercise_menu(frame, h, w)
                cv2.imshow("BioVision AI", frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord("q"):
                    break
                if key != 255 and chr(key) in EXERCISES:
                    current_exercise = EXERCISES[chr(key)]
                    rep_counter      = RepCounter(current_exercise)
                    pred_buf.clear(); conf_buf.clear()
                    session_log      = []
                    frame_n          = 0
                    print(f"Exercise: {current_exercise}  — SPACE or Q to stop")
                continue

            # ── Live analysis ──────────────────
            frame_n += 1
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            rgb.flags.writeable = False
            results = pose.process(rgb)
            rgb.flags.writeable = True

            angles    = None
            side_used = None
            is_good   = True
            issues    = []
            prediction  = 1
            confidence  = 0.5
            lms         = None

            if results.pose_landmarks:
                lms = results.pose_landmarks.landmark
                angles, side_used = best_angles(lms)

                if angles is not None:
                    # ML prediction
                    try:
                        ex_enc = encoder.transform([current_exercise])[0]
                    except ValueError:
                        ex_enc = 0

                    feat = scaler.transform([[
                        angles["Elbow_Angle"],
                        angles["Shoulder_Angle"],
                        angles["Hip_Angle"],
                        angles["Knee_Angle"],
                        ex_enc,
                    ]])
                    pred  = clf.predict(feat)[0]
                    proba = clf.predict_proba(feat)[0]
                    pred_buf.append(pred)
                    conf_buf.append(proba[pred])
                    prediction = int(round(np.mean(pred_buf)))
                    confidence = float(np.mean(conf_buf))

                    # Rule-based form check
                    is_good, issues = check_form(angles, current_exercise)

                    # Rep counting
                    rep_counter.update(angles)

                    # Log frame
                    session_log.append({
                        "Frame"          : frame_n,
                        "Exercise"       : current_exercise,
                        "Elbow_Angle"    : angles["Elbow_Angle"],
                        "Shoulder_Angle" : angles["Shoulder_Angle"],
                        "Hip_Angle"      : angles["Hip_Angle"],
                        "Knee_Angle"     : angles["Knee_Angle"],
                        "Form"           : "Good" if is_good else "Bad",
                        "ML_Prediction"  : "Good" if prediction == 1 else "Bad",
                        "Confidence_Pct" : round(confidence * 100, 1),
                        "Reps"           : rep_counter.count,
                        "Issues"         : "; ".join(r for _, r, *_ in issues),
                    })

                # Draw coloured skeleton
                draw_skeleton_colored(frame, lms, angles or {}, current_exercise, w, h)

            # FPS
            now      = time.time()
            fps      = 1.0 / max(now - prev_time, 1e-6)
            prev_time = now

            # Draw UI panel
            draw_left_panel(frame, current_exercise, is_good, confidence,
                            issues, angles, rep_counter, fps, h, side_used, lms, w)

            cv2.imshow("BioVision AI", frame)

            key = cv2.waitKey(1) & 0xFF
            if key in (ord("q"), ord(" ")):
                save_session(session_log, current_exercise)
                # Go back to exercise selection
                current_exercise = None
                rep_counter      = None
                pred_buf.clear(); conf_buf.clear()
            elif key == ord("s"):
                path = f"screenshots/shot_{screenshot_n:03d}.jpg"
                cv2.imwrite(path, frame)
                screenshot_n += 1
                print(f"Screenshot: {path}")

    cap.release()
    cv2.destroyAllWindows()
    print("Session ended.")


if __name__ == "__main__":
    main()
